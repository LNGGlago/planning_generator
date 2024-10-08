import yaml
from datetime import datetime, timedelta
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

# Charger les fichiers de configuration YAML
def load_config(file_path):
    with open(file_path, 'r') as file:
        return yaml.safe_load(file)

# Vérifie si une date est un jour férié, tombe pendant les vacances ou si la salle est indisponible
def is_holiday_vacation_or_unavailable(date, holidays, vacations, room_unavailability, room):
    date_str = str(date.date())
    if date_str in holidays:
        return True
    
    for vac in vacations:
        if vac['start'] <= date_str <= vac['end']:
            return True
    
    if room in room_unavailability:
        for period in room_unavailability[room]:
            if period['start'] <= date_str <= period['end']:
                return True
    
    return False

# Générer les dates pour une classe en excluant les jours fériés, les vacances et les indisponibilités de salle
def generate_dates(class_info, holidays, vacations, room_unavailability):
    day_mapping = {
        "Lundi": 0,
        "Mardi": 1,
        "Mercredi": 2,
        "Jeudi": 3,
        "Vendredi": 4,
        "Samedi": 5,
        "Dimanche": 6
    }
    class_day = day_mapping[class_info['day_of_week']]
    
    start_date = datetime.strptime("2024-09-01", "%Y-%m-%d")
    end_date = datetime.strptime("2025-08-31", "%Y-%m-%d")
    class_start_date = datetime.strptime(class_info.get('start_date', "2024-09-01"), "%Y-%m-%d")

    current_date = max(start_date, class_start_date)
    class_dates = []
    num_classes = class_info['num_classes']
    room = class_info['location']

    while len(class_dates) < num_classes and current_date <= end_date:
        if current_date.weekday() == class_day:
            if not is_holiday_vacation_or_unavailable(current_date, holidays, vacations, room_unavailability, room):
                class_dates.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)
    
    return class_dates

# Générer le planning avec les jours en abscisse et les classes en ordonnée
def generate_planning():
    classes = load_config('config/classes.yaml')['classes']
    holidays = load_config('config/holidays.yaml')['holidays']
    vacations = load_config('config/vacations.yaml')['vacations']
    room_unavailability = load_config('config/room_unavailability.yaml')['room_unavailability']
    
    # Regrouper les classes par lieux
    classes_by_location = defaultdict(list)
    for class_info in classes:
        classes_by_location[class_info['location']].append(class_info)
    
    # Créer un DataFrame avec les jours en colonnes
    start_date = datetime.strptime("2024-09-01", "%Y-%m-%d")
    end_date = datetime.strptime("2025-08-31", "%Y-%m-%d")
    dates = pd.date_range(start=start_date, end=end_date, freq='D')
    planning_df = pd.DataFrame(index=[], columns=dates)

    # Remplir le DataFrame avec les dates de cours et regrouper par lieu
    for location, classes in classes_by_location.items():
        planning_df.loc[location] = ""
        for class_info in classes:
            class_name = f"{class_info['name']} / {class_info['time']}"
            class_dates = generate_dates(class_info, holidays, vacations, room_unavailability)
            planning_df.loc[class_name] = ""
            for date in class_dates:
                planning_df.at[class_name, date] = "Cours"
    
    # Créer un fichier Excel et personnaliser l'affichage
    wb = Workbook()
    ws = wb.active

    # Créer les en-têtes d'années, mois, semaines, et jours
    ws.append(["Année"] + [date.strftime("%Y") for date in dates])
    ws.append(["Mois"] + [date.strftime("%B") for date in dates])
    ws.append(["Semaine"] + [f"Semaine {date.isocalendar()[1]}" for date in dates])
    ws.append(["Jour"] + [date.strftime("%d") for date in dates])

    # Remplir les lignes pour chaque classe en respectant les groupes par lieu
    row_num = 5
    for location, classes in classes_by_location.items():
        # Ajouter une ligne pour le lieu
        ws.append([location] + [""] * len(dates))
        row_num += 1

        for class_info in classes:
            class_name = f"{class_info['name']} / {class_info['time']}"
            row_data = [class_name] + [planning_df.at[class_name, date] for date in dates]
            ws.append(row_data)

            # Appliquer des styles : couleurs pour les cours, jours fériés et vacances
            for col in range(2, len(dates) + 2):
                cell_date = dates[col - 2].strftime("%Y-%m-%d")
                cell_value = ws.cell(row=row_num, column=col).value
                if cell_value == "Cours":
                    class_found = False
                    for cls in classes:
                        if cls['name'].strip() == class_info['name'].strip():
                            course_fill = PatternFill(start_color=cls['color'], end_color=cls['color'], fill_type="solid")
                            ws.cell(row=row_num, column=col).fill = course_fill
                            class_found = True
                            break
                    
                    if not class_found:
                        print(f"Erreur : Le cours '{class_info['name']}' n'a pas été trouvé dans la configuration.")

                if is_holiday_vacation_or_unavailable(datetime.strptime(cell_date, "%Y-%m-%d"), holidays, vacations, room_unavailability, location):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    ws.cell(row=row_num, column=col).font = Font(color="FFFFFF")
                    ws.cell(row=row_num, column=col).value = "Férié/Vac/Indisponibilité"

            row_num += 1

    # Fusionner les cellules pour les années
    current_year = ws.cell(row=1, column=2).value
    start_column = 2
    for col in range(2, len(dates) + 2):
        if ws.cell(row=1, column=col).value != current_year:
            ws.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=col-1)
            current_year = ws.cell(row=1, column=col).value
            start_column = col
    ws.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=len(dates) + 1)
    
    # Fusionner les cellules pour les mois
    current_month = ws.cell(row=2, column=2).value
    start_column = 2
    for col in range(2, len(dates) + 2):
        if ws.cell(row=2, column=col).value != current_month:
            ws.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=col-1)
            current_month = ws.cell(row=2, column=col).value
            start_column = col
    ws.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=len(dates) + 1)
    
    # Fusionner les cellules pour les semaines
    current_week = ws.cell(row=3, column=2).value
    start_column = 2
    for col in range(2, len(dates) + 2):
        if ws.cell(row=3, column=col).value != current_week:
            ws.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=col-1)
            current_week = ws.cell(row=3, column=col).value
            start_column = col
    ws.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=len(dates) + 1)

    output_file = os.path.join("output", "planning_formatted_with_years.xlsx")
    wb.save(output_file)
    print(f"Planning generated: {output_file}")

if __name__ == "__main__":
    os.makedirs("output", exist_ok=True)
    generate_planning()
